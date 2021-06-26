from dlink_home_product_spec import DLinkHomeProductSpec
from home_product_spec import TestSpec, ResultTestSpec, FullTestSpec, RegressionTestSpec, EasyTestSpec
from columns import FullTestColumns, RegressionTestColumns, EasyTestColumns

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from string import ascii_uppercase
from typing import Iterable, Dict, Tuple
from copy import copy
import json
import re



class ResultReport:

    __FileName = "testing_time_statistics.xlsx"

    @property
    def file_name(self):
        return self.__FileName


    @file_name.setter
    def file_name(self, filename):
        self.__FileName = filename


    @classmethod
    def general_report(cls, target_info):
        router_spec = DLinkHomeProductSpec(test_spec=ResultTestSpec(read_only=False))
        refer_sheet_page = router_spec.current_sheet_page
        workbook = Workbook()
        report_sheet_page = workbook.create_sheet(title="Test Time", index=0)
        print(f"[Debug] Target Info: {target_info}")
        """
        Note: (Not implement)
        Saving the content of data which be target to display and calculate testing time.
        Calculate the total testing time.
        """
        __testing_data = cls.testing_data_handling(target_info=target_info)
        cls.__copy_sheet_page(old_sheet_page=refer_sheet_page, new_sheet_page=report_sheet_page, target_value=__testing_data)
        cls.__merge_cells(sheet_page=report_sheet_page)
        return workbook


    @classmethod
    def testing_data_handling(cls, target_info: Dict[str, Dict]) -> Dict:
        spec = FullTestSpec()
        spec.loading_content()
        all_device_modules = spec.get_all_device_model()
        data_config = []

        print(f"target_info: {target_info}")
        print(f"target_info.values(): {target_info.values()}")
        for info in target_info.values():
            # Get value
            __device_model = info["device_model"]
            __spec_type = info["spec_type"]
            __items = info["items"]

            # Get the specific module index
            __device_model_excel_index = all_device_modules.index(__device_model)

            # Determine the spec type which need to use and calculate the testing time
            result_test_items, all_testing_items = cls.calculate_items_testing_time(spec_type=__spec_type, items=__items)

            # Save result into data-config
            data_config.append({
                "device_model": __device_model,
                "spec_type": __spec_type,
                "excel_index": __device_model_excel_index,
                "excel_columns": result_test_items,
                "excel_value": all_testing_items
            })

        # Add the NaN to the modules aren't target.
        list_data_index = 0
        final_data_config = []
        for index, __one_device_model in enumerate(all_device_modules):
            if index == 0:
                # the first time, get and check the data index
                __modul = data_config[list_data_index]["device_model"]
                __modul_spec_type = data_config[list_data_index]["spec_type"]
                __modul_excel_index = data_config[list_data_index]["excel_index"]
                __modul_excel_columns = data_config[list_data_index]["excel_columns"]
                __modul_excel_value = data_config[list_data_index]["excel_value"]
                list_data_index += 1

            if index != __modul_excel_index:
                # Insert default value into config
                final_data_config.append({
                    "device_model": __one_device_model,
                    "spec_type": None,
                    "excel_index": index,
                    "excel_columns": None,
                    "excel_value": None
                })
            else:
                # Insert the mapping index data into list.
                final_data_config.append({
                    "device_model": __modul,
                    "spec_type": __modul_spec_type,
                    "excel_index": __modul_excel_index,
                    "excel_columns": __modul_excel_columns,
                    "excel_value": __modul_excel_value
                })
                # Get the next value.
                if list_data_index < len(data_config):
                    __modul = data_config[list_data_index]["device_model"]
                    __modul_spec_type = data_config[list_data_index]["spec_type"]
                    __modul_excel_index = data_config[list_data_index]["excel_index"]
                    __modul_excel_columns = data_config[list_data_index]["excel_columns"]
                    __modul_excel_value = data_config[list_data_index]["excel_value"]
                    list_data_index += 1

        # print(f"Data config: {data_config}")
        return final_data_config


    @classmethod
    def calculate_items_testing_time(cls, spec_type: str, items: list) -> Tuple[list, Dict[str, float]]:
        """
        Sample Data:
        {
            "spec_testing_time_0":{
                "device_model":"",
                "spec_type":null,
                "items":[]
                },
            "spec_testing_time_1":{
                "device_model":"COVR-1103_A1",
                "spec_type":"full_test",
                "items":["{COVR-Reunion: 1}","{FOTA_In_Wizard: 12.5}","{Full_Total: =SUM(C2:R2)}"]
                }
        }
        """

        __testing_time = {}

        def calculate_time(result_items: list, item: dict):
            for __result_items in result_items:
                print(f"check value: {__result_items}")
                print(f"target value: {item}")
                print(f"target value: {item.keys()}")
                __key = list(item.keys())
                print(f"target value: {str(__key)}")
                print(f"target value: {str(__key[0])}")
                if re.search(re.escape(__result_items), str(list(item.keys())[0]), re.IGNORECASE):
                    __testing_time[str(item.keys[0])] = float(list(item.values())[0])
                    break
            else:
                __testing_time["Basic_Function"] = __testing_time.get("Basic_Function", 0) + float(list(item.values())[0])

        __result_spec = ResultTestSpec()
        final_items = None
        print(f"item: {items}")
        if spec_type == "full_test":
            final_items = __result_spec.get_full_test()
            print(f"final_items: {final_items}")
            for __item in items:
                print(f"__item: {__item}")
                print(f"__item type: {type(__item)}")
                calculate_time(result_items=final_items, item=dict(__item))
        elif spec_type == "regression_test":
            final_items = __result_spec.get_regression_test()
            for __item in items:
                calculate_time(result_items=final_items, item=dict(__item))
        elif spec_type == "sampling_test":
            final_items = __result_spec.get_easy_test()
            for __item in items:
                calculate_time(result_items=final_items, item=dict(__item))

        return final_items, __testing_time


    @classmethod
    def _get_ascii_char(cls, start_ascii: str = 0, end_ascii: str = 26) -> str:
        start_ascii_index = ascii_uppercase.index(start_ascii)
        end_ascii_index = ascii_uppercase.index(end_ascii)
        return ascii_uppercase[start_ascii_index:end_ascii_index+1]


    """
    Note:
    For coping value and style from one sheet page in excel to another sheet page in another excel,
    I refer to the answer in stackoverFlow:
    https://stackoverflow.com/questions/23332259/copy-cell-style-openpyxl


    Note:
    5nd row:  -->  Row
    C cell -->  model (5c start)
    D cell  -->  testing time value
    """


    @classmethod
    def __copy_sheet_page(cls, old_sheet_page: Worksheet, new_sheet_page: Worksheet, target_value: Iterable):
        for row_index, row in enumerate(old_sheet_page):
            # checksum = None
            for cell_index, cell in enumerate(row):
                if row_index < 4:
                    # new_cell = report_sheet_page.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                    new_cell = new_sheet_page[cell.coordinate]
                    cls.__copy_value(old_cell=cell, new_cell=new_cell)
                    cls.__copy_width(old_sheet_page=old_sheet_page, new_sheet_page=new_sheet_page, cell=cell)
                    cls.__copy_styles(old_cell=cell, new_cell=new_cell)
                else:
                    if cell_index == 2:
                        # new_cell = report_sheet_page.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                        new_cell = new_sheet_page[cell.coordinate]
                        cls.__copy_value(old_cell=cell, new_cell=new_cell)
                        cls.__copy_width(old_sheet_page=old_sheet_page, new_sheet_page=new_sheet_page, cell=cell)
                        cls.__copy_styles(old_cell=cell, new_cell=new_cell)
                        # checksum = cls.chk_device_module(current=cell, target=target_value)
                    elif cell_index < 2:
                        # new_cell = report_sheet_page.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                        new_cell = new_sheet_page[cell.coordinate]
                        cls.__copy_value(old_cell=cell, new_cell=new_cell)
                        cls.__copy_width(old_sheet_page=old_sheet_page, new_sheet_page=new_sheet_page, cell=cell)
                        cls.__copy_styles(old_cell=cell, new_cell=new_cell)
                    else:
                        # The target detail info about testing time TE needs recording.
                        if (row_index - 4) < len(target_value):
                            __current_modul_info = target_value[row_index - 4]
                            __current_modul_excel_val = __current_modul_info["excel_value"]
                            if __current_modul_excel_val:
                                __testing_item_val = 0

                                if cell_index == FullTestColumns.BasicFunction:
                                    __testing_item_val = __current_modul_excel_val.get("Basic_Function", 0)
                                if cell_index == FullTestColumns.ParentalControl:
                                    __testing_item_val = __current_modul_excel_val.get("Parental_control", 0)
                                if cell_index == FullTestColumns.AdvanceParentalControl:
                                    __testing_item_val = __current_modul_excel_val.get("Advance_parental_control", 0)
                                if cell_index == FullTestColumns.DLinkDeFendFullFunction:
                                    __testing_item_val = __current_modul_excel_val.get("D-Link_DeFend_full_function", 0)
                                if cell_index == FullTestColumns.FOTAInWizard:
                                    __testing_item_val = __current_modul_excel_val.get("FOTA_In_Wizard", 0)
                                if cell_index == FullTestColumns.RemoteManagement:
                                    __testing_item_val = __current_modul_excel_val.get("Remote_management", 0)
                                if cell_index == FullTestColumns.GoogleAndAlexaTest:
                                    __testing_item_val = __current_modul_excel_val.get("Google_and_Alexa_Test", 0)
                                
                                if cell_index == RegressionTestColumns.BasicFunction:
                                    __testing_item_val = __current_modul_excel_val.get("Basic_Function", 0)
                                if cell_index == RegressionTestColumns.DLinkDeFendRegressionFunction:
                                    __testing_item_val = __current_modul_excel_val.get("D-Link_DeFend_full_function", 0)
                                if cell_index == RegressionTestColumns.FOTAInWizard:
                                    __testing_item_val = __current_modul_excel_val.get("FOTA_in_wizard", 0)
                                if cell_index == RegressionTestColumns.RemoteManagement:
                                    __testing_item_val = __current_modul_excel_val.get("Remote_management", 0)
                                
                                if cell_index == EasyTestColumns.EasyTest:
                                    __testing_item_val = __current_modul_excel_val.get("Easy_Test", 0)

                                new_cell = new_sheet_page[cell.coordinate]
                                cls.__insert_value(new_cell=new_cell, value=__testing_item_val)
                                print("Do something to update value")
                            else:
                                new_cell = new_sheet_page[cell.coordinate]
                                cls.__insert_value(new_cell=new_cell, value=0)
                                cls.__copy_width(old_sheet_page=old_sheet_page, new_sheet_page=new_sheet_page, cell=cell)
                                cls.__copy_styles(old_cell=cell, new_cell=new_cell)
                                # print("Insert 0 to the column.")
                        else:
                            # print("Insert 0 to the column.")
                            pass


    @classmethod
    def chk_device_module(cls, current: Cell, target: Dict) -> bool:
        """
        Think about procedure.
        """
        print(f"target: {target}")
        if current.value is not None:
            print(f"[DEBUG] current value: {current.value}")
        for device_info_index, device_info in target.items():
            if current.value == device_info["device_model"]:
                return True
        else:
            return False


    @classmethod
    def __insert_value(cls, new_cell: Cell, value) -> None:
        new_cell.value = value


    @classmethod
    def __copy_value(cls, old_cell: Cell, new_cell: Cell) -> None:
        new_cell.value = old_cell.value


    @classmethod
    def __set_value(cls, new_cell: Cell, value: int) -> None:
        pass


    @classmethod
    def __copy_width(cls, old_sheet_page: Worksheet, new_sheet_page: Worksheet, cell: Cell) -> None:
        new_sheet_page.column_dimensions[get_column_letter(cell.column)].width = old_sheet_page.column_dimensions[get_column_letter(cell.column)].width


    @classmethod
    def __copy_styles(cls, old_cell, new_cell) -> None:
        if old_cell.has_style:
            new_cell.font = copy(old_cell.font)
            new_cell.border = copy(old_cell.border)
            new_cell.fill = copy(old_cell.fill)
            new_cell.number_format = copy(old_cell.number_format)
            new_cell.protection = copy(old_cell.protection)
            new_cell.alignment = copy(old_cell.alignment)


    @classmethod
    def __merge_cells(cls, sheet_page: Worksheet):
        # For topic
        sheet_page.merge_cells(range_string="A1:J1")
        sheet_page.merge_cells(range_string="A2:J2")
        sheet_page.merge_cells(range_string="K1:P2")

        # For Series column
        sheet_page.merge_cells(range_string="A3:B4")
        # For Model column
        sheet_page.merge_cells(range_string="C3:C4")

        # For Full test topic
        sheet_page.merge_cells(range_string="D3:J3")
        # For Regression test topic
        sheet_page.merge_cells(range_string="K3:N3")
        # For Easy test topic
        sheet_page.merge_cells(range_string="O3:O4")
        # For Select test topic
        sheet_page.merge_cells(range_string="P3:P4")

        # For DIR series
        sheet_page.merge_cells(range_string="B5:B20")
        # For EXO series
        sheet_page.merge_cells(range_string="B21:B28")
        # For Easy mesh series
        sheet_page.merge_cells(range_string="B29:B42")
        # For COVR series
        sheet_page.merge_cells(range_string="B43:B50")
        # For DAP series
        sheet_page.merge_cells(range_string="B51:B63")

        # For Summary 1
        sheet_page.merge_cells(range_string="A64:O64")
        # For Summary 2
        sheet_page.merge_cells(range_string="K65:O65")

