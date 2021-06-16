from openpyxl import Workbook, load_workbook
from string import ascii_uppercase
from typing import List, Dict


class ReadSpec:

    __DLink_Test_Time_Spec_File_Name: str = "D-Link_Wi-Fi_config_time_20201219.xlsx"
    __First_Data_Row_Index: int = 5
    __Last_Data_Row_Index: int = 63

    __Device_Series_Column: str = "B"
    __Device_Models_Column: str = "C"
    __Full_Test_Case: Dict[str, str] = {
        "start_column": "D",
        "end_column": "J",
    }
    __Regression_Test_Case: Dict[str, str] = {
        "start_column": "K",
        "end_column": "N",
    }
    __Easy_Test_Column: str = "O"
    __Sampling_Test_Column: str = ""

    __DLink_Spec_WorkBook: Workbook = None
    __DLink_Spec_WorkBook_Sheet_Page: Workbook = None


    def dev_and_test_main(self) -> None:
        self.loading_content()

        # # Change the sheet page to target we want.
        # Method 1.
        # dlink_spec_sheet_names = dlink_spec_wb.sheetnames
        # sheet_page = dlink_spec_wb[dlink_spec_sheet_names[1]]
        # Method 2.
        self.__DLink_Spec_WorkBook_Sheet_Page = self.__DLink_Spec_WorkBook.get_sheet_by_name(name="group test time")

        device_models = self.get_device_models()
        print(device_models)

        all_test_cases = self.get_all_test_cases()
        print(all_test_cases)

        self.close_workbook()


    def loading_content(self) -> None:
        self.__DLink_Spec_WorkBook = load_workbook(filename=self.__DLink_Test_Time_Spec_File_Name, read_only=True)


    def get_sheet_by_name(self, name: str) -> None:
        self.__DLink_Spec_WorkBook_Sheet_Page = self.__DLink_Spec_WorkBook.get_sheet_by_name(name=name)


    def get_device_series(self) -> List[str]:
        return self.__get_all_row_data_with_column(sheet_page=self.__DLink_Spec_WorkBook_Sheet_Page, column=self.__Device_Series_Column)


    def get_device_models(self) -> List[str]:
        return self.__get_all_row_data_with_column(sheet_page=self.__DLink_Spec_WorkBook_Sheet_Page, column=self.__Device_Models_Column)


    def get_all_test_cases(self) -> List[str]:
        return self.__get_all_row_data_with_row(sheet_page=self.__DLink_Spec_WorkBook_Sheet_Page,
                                                row=4,
                                                columns=self.__get_ascii_char(start_ascii=self.__Full_Test_Case["start_column"],
                                                                              end_ascii=self.__Regression_Test_Case["end_column"]))


    def get_full_test_cases(self) -> List[str]:
        start_column = self.__Full_Test_Case["start_column"]
        end_column = self.__Full_Test_Case["end_column"]
        return self.__get_all_row_data_with_row(sheet_page=self.__DLink_Spec_WorkBook_Sheet_Page,
                                                row=4,
                                                columns=self.__get_ascii_char(start_ascii=start_column,
                                                                              end_ascii=end_column))


    def is_one_of_full_test(self, test_case: str) -> bool:
        return test_case in self.get_full_test_cases()


    def get_regression_test_cases(self) -> List[str]:
        start_column = self.__Full_Test_Case["start_column"]
        end_column = self.__Full_Test_Case["end_column"]
        return self.__get_all_row_data_with_row(sheet_page=self.__DLink_Spec_WorkBook_Sheet_Page,
                                                row=4,
                                                columns=self.__get_ascii_char(start_ascii=start_column,
                                                                              end_ascii=end_column))


    def is_one_of_regression_test(self, test_case: str) -> bool:
        return test_case in self.get_regression_test_cases()


    def get_easy_test(self) -> List[str]:
        return self.__get_all_row_data_with_column(sheet_page=self.__DLink_Spec_WorkBook_Sheet_Page, column=self.__Easy_Test_Column)


    def get_sampling_test(self) -> List[str]:
        return self.__get_all_row_data_with_column(sheet_page=self.__DLink_Spec_WorkBook_Sheet_Page, column=self.__Easy_Test_Column)


    def __get_ascii_char(self, start_ascii: str = 0, end_ascii: str = 26) -> str:
        start_ascii_index = ascii_uppercase.index(start_ascii)
        end_ascii_index = ascii_uppercase.index(end_ascii)
        return ascii_uppercase[start_ascii_index:end_ascii_index+1]


    def __get_all_row_data_with_row(self, sheet_page, row, columns) -> List[str]:
        result_data: List[str] = []
        for column in columns:
            cell_value = sheet_page[f"{column}{row}"].value  # the value of the specific cell
            if cell_value is not None:
                result_data.append(cell_value)
        return result_data


    def __get_all_row_data_with_column(self, sheet_page, column) -> List[str]:
        result_data: List[str] = []
        for row in range(self.__First_Data_Row_Index, self.__Last_Data_Row_Index + 1):
            cell_value = sheet_page[f"{column}{row}"].value  # the value of the specific cell
            if cell_value is not None:
                result_data.append(cell_value)
        return result_data


    def __get_all_row_data_with_columns(self, sheet_page, columns) -> List[List[str]]:
        result_data: List[List[str]] = []
        for row in range(self.__First_Data_Row_Index, self.__Last_Data_Row_Index + 1):
            data_row: List[str] = []
            for column in columns:
                cell_value = sheet_page[f"{column}{row}"].value  # the value of the specific cell
                if cell_value is not None:
                    data_row.append(cell_value)
            result_data.append(data_row)
        return result_data


    def close_workbook(self):
        self.__DLink_Spec_WorkBook.close()


# if __name__ == '__main__':
#
#     read_spec = ReadSpec()
#     read_spec.dev_and_test_main()
