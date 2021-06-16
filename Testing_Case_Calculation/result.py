from dlink_home_product_spec import DLinkHomeProductSpec
from home_product_spec import TestSpec, ResultTestSpec, FullTestSpec, RegressionTestSpec, EasyTestSpec

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from string import ascii_uppercase
from typing import Iterable, Dict
from copy import copy
import json



class ResultReport:

    __FileName = "testing_time_statistics.xlsx"

    @property
    def file_name(self):
        return self.__FileName


    @file_name.setter
    def file_name(self, filename):
        self.__FileName = filename


    @classmethod
    def general_report(cls, target_info):
        router_spec = DLinkHomeProductSpec(test_spec=ResultTestSpec(read_only=False))
        refer_sheet_page = router_spec.current_sheet_page
        workbook = Workbook()
        report_sheet_page = workbook.create_sheet(title="Test Time", index=0)
        print(f"[Debug] Target Info: {target_info}")
        """
        Note: (Not implement)
        Saving the content of data which be target to display and calculate testing time.
        Calculate the total testing time.
        """
        cls.__copy_sheet_page(old_sheet_page=refer_sheet_page, new_sheet_page=report_sheet_page, target_value=target_info)
        cls.__merge_cells(sheet_page=report_sheet_page)
        return workbook


    @classmethod
    def _get_ascii_char(cls, start_ascii: str = 0, end_ascii: str = 26) -> str:
        start_ascii_index = ascii_uppercase.index(start_ascii)
        end_ascii_index = ascii_uppercase.index(end_ascii)
        return ascii_uppercase[start_ascii_index:end_ascii_index+1]


    """
    Note:
    For coping value and style from one sheet page in excel to another sheet page in another excel,
    I refer to the answer in stackoverFlow:
    https://stackoverflow.com/questions/23332259/copy-cell-style-openpyxl


    Note:
    5nd row:  -->  Row
    C cell -->  model (5c start)
    D cell  -->  testing time value
    """


    @classmethod
    def __copy_sheet_page(cls, old_sheet_page: Worksheet, new_sheet_page: Worksheet, target_value: Iterable):
        for row_index, row in enumerate(old_sheet_page):
            checksum = None
            for cell_index, cell in enumerate(row):
                if row_index < 4:
                    # new_cell = report_sheet_page.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                    new_cell = new_sheet_page[cell.coordinate]
                    cls.__copy_value(old_cell=cell, new_cell=new_cell)
                    cls.__copy_width(old_sheet_page=old_sheet_page, new_sheet_page=new_sheet_page, cell=cell)
                    cls.__copy_styles(old_cell=cell, new_cell=new_cell)
                else:
                    if cell_index == 2:
                        # new_cell = report_sheet_page.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                        new_cell = new_sheet_page[cell.coordinate]
                        cls.__copy_value(old_cell=cell, new_cell=new_cell)
                        cls.__copy_width(old_sheet_page=old_sheet_page, new_sheet_page=new_sheet_page, cell=cell)
                        cls.__copy_styles(old_cell=cell, new_cell=new_cell)
                        checksum = cls.chk_device_module(current=cell, target=json.loads(target_value))
                    elif cell_index < 2:
                        # new_cell = report_sheet_page.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                        new_cell = new_sheet_page[cell.coordinate]
                        cls.__copy_value(old_cell=cell, new_cell=new_cell)
                        cls.__copy_width(old_sheet_page=old_sheet_page, new_sheet_page=new_sheet_page, cell=cell)
                        cls.__copy_styles(old_cell=cell, new_cell=new_cell)
                    else:
                        if checksum is True:
                            print("Do something to update value")
                        else:
                            pass
                            # print("Just insert empty value.")



    @classmethod
    def chk_device_module(cls, current: Cell, target: Dict) -> bool:
        if current.value is not None:
            print(f"[DEBUG] current value: {current.value}")
        for device_info_index, device_info in target.items():
            if current.value == device_info["device_model"]:
                return True
        else:
            return False


    @classmethod
    def __copy_value(cls, old_cell: Cell, new_cell: Cell) -> None:
        new_cell.value = old_cell.value


    @classmethod
    def __set_value(cls, new_cell: Cell, value: int) -> None:
        pass


    @classmethod
    def __copy_width(cls, old_sheet_page: Worksheet, new_sheet_page: Worksheet, cell: Cell) -> None:
        new_sheet_page.column_dimensions[get_column_letter(cell.column)].width = old_sheet_page.column_dimensions[get_column_letter(cell.column)].width


    @classmethod
    def __copy_styles(cls, old_cell, new_cell) -> None:
        if old_cell.has_style:
            new_cell.font = copy(old_cell.font)
            new_cell.border = copy(old_cell.border)
            new_cell.fill = copy(old_cell.fill)
            new_cell.number_format = copy(old_cell.number_format)
            new_cell.protection = copy(old_cell.protection)
            new_cell.alignment = copy(old_cell.alignment)


    @classmethod
    def __merge_cells(cls, sheet_page: Worksheet):
        # For topic
        sheet_page.merge_cells(range_string="A1:J1")
        sheet_page.merge_cells(range_string="A2:J2")
        sheet_page.merge_cells(range_string="K1:P2")

        # For Series column
        sheet_page.merge_cells(range_string="A3:B4")
        # For Model column
        sheet_page.merge_cells(range_string="C3:C4")

        # For Full test topic
        sheet_page.merge_cells(range_string="D3:J3")
        # For Regression test topic
        sheet_page.merge_cells(range_string="K3:N3")
        # For Easy test topic
        sheet_page.merge_cells(range_string="O3:O4")
        # For Select test topic
        sheet_page.merge_cells(range_string="P3:P4")

        # For DIR series
        sheet_page.merge_cells(range_string="B5:B20")
        # For EXO series
        sheet_page.merge_cells(range_string="B21:B28")
        # For Easy mesh series
        sheet_page.merge_cells(range_string="B29:B42")
        # For COVR series
        sheet_page.merge_cells(range_string="B43:B50")
        # For DAP series
        sheet_page.merge_cells(range_string="B51:B63")

        # For Summary 1
        sheet_page.merge_cells(range_string="A64:O64")
        # For Summary 2
        sheet_page.merge_cells(range_string="K65:O65")

