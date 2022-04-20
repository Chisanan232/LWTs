from src import create_app

from openpyxl import load_workbook, Workbook
from flask.testing import FlaskClient, FlaskCliRunner
from flask import Flask
from io import BytesIO
import traceback
import pytest
import ast


@pytest.fixture()
def app() -> Flask:
    _app = create_app()
    _app.config["debug"] = True
    # _app.run(host="0.0.0.0", port=5224, debug=True)
    yield _app


@pytest.fixture()
def client(app: Flask) -> FlaskClient:
    return app.test_client()


@pytest.fixture()
def runner(app: Flask) -> FlaskCliRunner:
    return app.test_cli_runner()



class TestSpecAPIs:

    _Decoding = "ascii"

    def test_root_api(self, client: FlaskClient):
        """
        Testing Info:
          API: '/'
          HTTP Method: GET
          Parameters: None
        
        API Description: 
          Only for testing to ensure that the service is activated.
        """

        _api = "/"
        _response = client.get(_api)
        _data = _response.data.decode(self._Decoding)

        assert _data == "Welcom to Leslie Wish Tools - 1", "The web message should be equal to the expected characters."


    def test_device_modules(self, client: FlaskClient):
        """
        Testing Info:
          API: '/api/v1/deviceModels'
          HTTP Method: GET
          Parameters: None
        
        API Description: 
          Get all the device modules. This API be used by the selector at option 'Device Model' in Font-End.
        """

        _api = "/api/v1/deviceModels"
        _response = client.get(_api)
        _data = _response.data.decode(self._Decoding)

        try:
            _parsed_data = ast.literal_eval(_data)
        except Exception as e:
            assert False, "It should not occur any error when it parsed data."
        else:
            assert type(_parsed_data) is list, "The return value should be a list type value."


    def test_get_test_items_with_device_modules(self, client: FlaskClient):
        """
        Testing Info:
          API: '/api/v1/testing-time'
          HTTP Method: POST
          Parameters: 
            * spec: Test plan
            * deviceModel: Device module 
        
        API Description: 
          Get the details about how long TE (Test Engineer) needs to take.
        """

        spec_type = "full"
        device_model = "DIR-882"

        _api = "/api/v1/testing-time"
        _response = client.post(
            _api, 
            data={
                "spec": spec_type, 
                "deviceModel": device_model}
            )
        _data = _response.data.decode(self._Decoding)

        try:
            _parsed_data = ast.literal_eval(_data)
        except Exception as e:
            assert False, "It should not occur any error when it parsed data."
        else:
            assert type(_parsed_data) is dict, "The value should be a Dict type data."


    def test_export_test_time_result(self, client: FlaskClient):
        """
        Testing Info:
          API: '/api/v1/export-testing-time-result'
          HTTP Method: POST
          Parameters: 
            * target: The JSON type (that's in JavaScript, it's Dict type in Python) value which saves all testing info (includes device modules, test plan, testing time).
        
        API Description: 
          Export all the testing info (includes device modules, test plan, testing time) as xlsx format file.
        """

        # # # # Prepare the body (the parameters of API) 
        spec_type = "full_test"
        device_model = "DIR-882"

        # _test_item_info_char = {"Advance parental control":0,"COVR-Reunion":0,"D-Link DeFend full function":0,"FOTA In Wizard":0,"Internet":2.1,"Management-AP":0,"Management-router":3,"Operation":0,"Parental control":2,"Remote management":0,"Side Menu":0.5,"main page-router":1.4,"main page-AP":0,"mesh compatibility":0,"smart connect Wi-Fi Settings":2,"wizard":3.9}
        # _test_item_info_str = json.dumps(_test_item_info_char).encode("utf-8")
        _test_item_info_char = ["{\"Internet\": 2.1}","{\"Management-router\": 3}"]
        _test_item = {
            "device_model": device_model,
            "spec_type": spec_type,
            "items": _test_item_info_char
        }
        _test_result = {"spec_testing_time_0": _test_item}

        _api = "/api/v1/export-testing-time-result"
        _response = client.post(
            _api, 
            data={"target": str(_test_result)}
            # follow_redirects=True,
            # content_type="application/json"
            )
        _data = _response.data
        _workbook: Workbook = load_workbook(BytesIO(_data))
        _worksheet = _workbook.worksheets[0]
        _workersheet_iter_rows = _worksheet.iter_rows()
        for _rows in _workersheet_iter_rows:
            _device_module = _rows[2].value
            if _device_module == device_model:
                assert _rows[3].value == 5.1, "The value of cell should be equal to the data we input into API."

